#!/usr/bin/env python3
"""Build the visual + spreadsheet report for a content-gap run.

Reads a run directory (meta.json, clusters.json, gaps.json, our.json,
competitor-*.json) and produces:

  report.html  — a rich, self-contained page (no CDN): KPI cards, cluster
                 coverage diagram, per-cluster SIDE-BY-SIDE of what each page
                 actually wrote, content-similarity scoring, per-page H1/H2/H3
                 lists, internal/external link tables, image lists, ranking view,
                 external-brand mentions. Every section deep-links to the live
                 page (Chrome text-fragment) and has an in-report anchor.
  report.xlsx  — every piece of content in filterable sheets: Full Content,
                 Section Comparison, Similarity, H1/H2/H3, Heading Coverage,
                 Content Coverage, FAQ Coverage,
                 Heading Markup Gaps (topics covered but not in any H-tag),
                 Structure Fix (issues + corrected reading order), Internal/External
                 Links, Images, Ranking, External Brands, Gaps, FAQ, Quality.
                 (CSV fallback if openpyxl is missing.)

Standard library only for the HTML (similarity uses difflib). XLSX needs openpyxl.
Usage: python build_report.py <run_dir>
"""
import csv
import difflib
import html
import json
import os
import re
import sys
from urllib.parse import quote


# ----------------------------- load -----------------------------------------
def _load(path, default):
    try:
        with open(path, "r", encoding="utf-8-sig") as fh:
            return json.load(fh)
    except Exception:
        return default


def load_run(run_dir):
    meta = _load(os.path.join(run_dir, "meta.json"), {})
    clusters = _load(os.path.join(run_dir, "clusters.json"), {"clusters": []})
    gaps = _load(os.path.join(run_dir, "gaps.json"), {})
    pages = []
    for name in sorted(os.listdir(run_dir)):
        if name.endswith(".json") and (name == "our.json" or name.startswith("competitor")):
            pages.append(_load(os.path.join(run_dir, name), {}))
    return meta, clusters, gaps, pages


def brand_order(meta, gaps, pages):
    your = gaps.get("your_brand") or meta.get("your_brand") or "OUR PAGE"
    order = [your]
    for p in pages:
        b = p.get("brand")
        if b and b not in order:
            order.append(b)
    return your, order


def esc(x):
    return html.escape(str(x if x is not None else ""))


# Strip form-widget noise that leaks into extracted section text — long runs of
# phone country codes (e.g. "+91 +1 (USA) +1 (CAN) +61 (AUS) +65 +962 ...") from
# the phone-number dropdown on premium-calculator forms. 4+ consecutive "+code"
# tokens is a country picker, never prose.
_PHONE_CODE_RUN = re.compile(r'(?:\+\d{1,4}(?:\s*\([^)]{1,8}\))?\s*){4,}')
# Income-band picker runs, e.g. "< 2.5 Lakhs 2.5 - 5 Lakhs 5 - 7.5 Lakhs ...".
_INCOME_BAND_RUN = re.compile(
    r'(?:[<>]?\s*\d[\d.,]*\s*(?:-\s*\d[\d.,]*)?\s*Lakhs?\b[\s,]*){2,}', re.I)


def clean_text(t):
    if not t:
        return t
    t = _PHONE_CODE_RUN.sub(" ", t)
    t = _INCOME_BAND_RUN.sub(" ", t)
    return re.sub(r"\s{2,}", " ", t).strip()


def page_by_brand(pages):
    return {p.get("brand"): p for p in pages}


# --------------------- resilient brand-key resolution -----------------------
# clusters.json / gaps.json are written by the gap-analyst agent and SHOULD key
# every brand map under the exact canonical brand string from meta.json. If the
# agent drifts (short form, "OUR PAGE", the URL), an exact lookup silently zeroes
# the column. Resolve case/whitespace-insensitively, and never fail silently.
def _norm(s):
    return re.sub(r"\s+", " ", str(s if s is not None else "").strip().lower())


def resolve_brand_key(brand_map, canonical):
    """Return the key in brand_map matching `canonical` — exact first, then
    case/whitespace-insensitive. None if nothing matches."""
    if not isinstance(brand_map, dict) or not brand_map:
        return None
    if canonical in brand_map:
        return canonical
    target = _norm(canonical)
    for k in brand_map:
        if _norm(k) == target:
            return k
    return None


def map_get(brand_map, canonical, default=None):
    """brand_map value for the canonical brand via resilient resolution."""
    key = resolve_brand_key(brand_map or {}, canonical)
    return (brand_map or {}).get(key, default) if key is not None else default


def brand_entry(brand_map, canonical):
    """The dict entry for a canonical brand, or {} — for cluster brand maps."""
    return map_get(brand_map, canonical, {}) or {}


# Internal links inside nav / header / footer are site-wide boilerplate, not
# on-page editorial links. Exclude them everywhere (counts, charts, tables, gap
# analysis) so internal-link comparisons reflect real in-content linking only.
NAV_FOOTER_SCOPES = {"nav", "header", "footer"}


def onpage_internal_links(page):
    """A page's internal links with nav/header/footer boilerplate stripped out."""
    return [l for l in (page.get("internal_links") or [])
            if (l.get("scope") or "in-content") not in NAV_FOOTER_SCOPES]


def covers(info):
    """True if a brand actually addresses a cluster (present flag or depth > 0)."""
    return bool(info.get("present")) or int(info.get("depth", 0) or 0) > 0


def unique_clusters_by_brand(clusters, order):
    """Map brand -> [cluster names] that ONLY that brand covers — the unique
    content angle no other compared page has."""
    out = {b: [] for b in order}
    for c in clusters.get("clusters", []):
        owners = [b for b in order if covers(brand_entry(c.get("brands"), b))]
        if len(owners) == 1:
            out[owners[0]].append(c.get("name"))
    return out


def _normh(t):
    """Normalise a heading for cross-company matching (case/punctuation-blind)."""
    t = re.sub(r"[^\w\s]", " ", (t or "").lower())
    return re.sub(r"\s+", " ", t).strip()


def heading_coverage(pages, order):
    """Every H1/H2/H3 heading across all pages, as a NESTED tree in document
    order: each H1 is followed by its child H2s, each H2 by its child H3s
    (the 'Term Category Page' hierarchy view). Headings are matched
    case/punctuation-insensitively within a level. The tree is anchored to the
    first page (your page) and competitor-unique branches are appended where
    they nest. Returns rows in depth-first order: [{level, text, present:{brand:bool}}].
    """
    pbb = page_by_brand(pages)
    nodes = {}   # (level, norm) -> node
    root = []    # top-level H1 nodes, in first-seen order

    def node(level, nk, text):
        k = (level, nk)
        n = nodes.get(k)
        if n is None:
            n = {"level": level, "text": text, "present": set(),
                 "children": [], "attached": False}
            nodes[k] = n
        return n

    for b in order:
        p = pbb.get(b) or {}
        cur1 = cur2 = None  # keys of the current H1 / H2 while walking this page
        for h in (p.get("heading_outline") or []):
            lvl = int(h.get("level", 0) or 0)
            if lvl not in (1, 2, 3):
                continue
            t = (h.get("text") or "").strip()
            if not t:
                continue
            nk = _normh(t)
            n = node(lvl, nk, t)
            n["present"].add(b)
            if not n["attached"]:
                n["attached"] = True
                parent = None
                if lvl == 2:
                    parent = cur1
                elif lvl == 3:
                    parent = cur2 or cur1
                if parent is not None and parent in nodes:
                    nodes[parent]["children"].append(n)
                else:
                    root.append(n)
            if lvl == 1:
                cur1, cur2 = (1, nk), None
            elif lvl == 2:
                cur2 = (2, nk)

    rows = []

    def dfs(n):
        rows.append({"level": n["level"], "text": n["text"],
                     "present": {b: (b in n["present"]) for b in order}})
        for ch in n["children"]:
            dfs(ch)

    for n in root:
        dfs(n)
    return rows


def covered_not_in_headings(page):
    """Pseudo-headings on this page whose text is NOT any real H1/H2/H3 — topics
    the page presents as section titles but ships inside <p>/<div>/<span> (or a
    styled class), so a crawler / AI bot never registers them as headings.
    Returns rows in document order with a suggested heading level."""
    real = set()
    for h in (page.get("heading_outline") or []):
        if int(h.get("level", 0) or 0) in (1, 2, 3):
            real.add(_normh(h.get("text")))
    seen, out = set(), []
    for ph in (page.get("pseudo_headings") or []):
        t = (ph.get("text") or "").strip()
        nk = _normh(t)
        if not nk or nk in real or nk in seen:
            continue
        seen.add(nk)
        pl = int(ph.get("parent_level", 0) or 0)
        out.append({"text": t, "tag": ph.get("tag") or "p",
                    "class": (ph.get("class") or "").split()[0] if ph.get("class") else "",
                    "parent_heading": ph.get("parent_heading") or "",
                    "parent_level": pl, "seq": int(ph.get("seq", 0) or 0),
                    "suggested": "H3" if pl >= 2 else "H2"})
    return out


def content_coverage(clusters, order):
    """Every topic cluster across the pages as a flat Yes/No matrix: does each
    company cover this topic ANYWHERE on the page — in a heading OR in body copy,
    even if it is worded differently? This is the header-blind companion to
    heading_coverage(): that one asks 'is the topic an H-tag?', this asks 'is the
    topic present at all?'. Driven by the gap-analyst's block-to-block clustering,
    which already aligns differently-phrased coverage of the same topic, so a
    topic everyone covers but writes up in their own words still reads as Yes.
    Returns [{topic, present:{brand:bool}}] in cluster order."""
    rows = []
    for c in clusters.get("clusters", []):
        present = {b: covers(brand_entry(c.get("brands"), b)) for b in order}
        rows.append({"topic": c.get("name") or "", "present": present})
    return rows


_FAQ_STOP = set((
    "a an the is are was were do does did of to for in on at with my our your their i you it its "
    "what which how when why who whom whose can could will would should shall may might must need "
    "and or vs versus be been being have has had this that these those if then than as about into "
    "by from get got give vs. e g eg etc"
).split())


def _faq_tokens(q):
    """Content tokens of an FAQ question, stop-words stripped — used to group
    differently-worded questions about the same thing."""
    toks = re.sub(r"[^\w\s]", " ", (q or "").lower()).split()
    return set(t for t in toks if t not in _FAQ_STOP and len(t) > 2)


def faq_coverage(gaps, pages, order, your):
    """Every distinct FAQ across the pages as a Yes/No matrix: does each company
    answer this question — even if they phrase it differently? Differently-worded
    questions about the same thing are grouped by content-token overlap so e.g.
    'Is a medical test required?' and 'Do I need a medical test for term
    insurance?' count as one row. Prefers an explicit gaps['faq_coverage'] from
    the gap-analyst (semantic, most accurate); otherwise clusters the raw page
    FAQs deterministically. Returns [{question, present:{brand:bool}}],
    most-covered first."""
    # 1) gap-analyst semantic enrichment, if provided
    enr = gaps.get("faq_coverage")
    if isinstance(enr, list) and enr:
        rows = []
        for e in enr:
            if not isinstance(e, dict):
                continue
            q = e.get("question") or e.get("q") or ""
            pres = e.get("present")
            if isinstance(pres, dict):
                present = {b: bool(map_get(pres, b, False)) for b in order}
            else:
                ab = e.get("answered_by") or []
                present = {b: any(_norm(b) == _norm(x) for x in ab) for b in order}
            if q:
                rows.append({"question": q, "present": present})
        if rows:
            rows.sort(key=lambda r: sum(r["present"].values()), reverse=True)
            return rows
    # 2) deterministic fallback — cluster page FAQs by token overlap
    pbb = page_by_brand(pages)
    groups = []  # [{rep, has_ours, toks, present:set}]
    for b in order:
        p = pbb.get(b) or {}
        for fq in (p.get("faqs") or []):
            q = (fq.get("question") or "").strip()
            tk = _faq_tokens(q)
            if not q or not tk:
                continue
            best, bestj = None, 0.0
            for g in groups:
                inter = len(tk & g["toks"])
                union = len(tk | g["toks"]) or 1
                j = inter / union
                if j > bestj:
                    best, bestj = g, j
            if best is not None and bestj >= 0.5:
                best["present"].add(b)
                best["toks"] |= tk
                # prefer OUR phrasing as the representative, else the shortest
                if (b == your and not best["has_ours"]) or \
                   (best["has_ours"] == (b == your) and len(q) < len(best["rep"])):
                    best["rep"] = q
                if b == your:
                    best["has_ours"] = True
            else:
                groups.append({"rep": q, "has_ours": (b == your),
                               "toks": set(tk), "present": {b}})
    rows = [{"question": g["rep"], "present": {b: (b in g["present"]) for b in order}}
            for g in groups]
    rows.sort(key=lambda r: sum(r["present"].values()), reverse=True)
    return rows


def merged_outline(page):
    """The page's real headings + its not-in-heading topics, interleaved in true
    document order (by seq). Real headings carry their current tag; pseudo rows
    are flagged with the tag they SHOULD become. This is the 'corrected reading
    order' a crawler should see."""
    rows = []
    for h in (page.get("heading_outline") or []):
        lvl = int(h.get("level", 0) or 0)
        if lvl in (1, 2, 3):
            rows.append({"seq": int(h.get("seq", 0) or 0), "kind": "real",
                         "tag": "H%d" % lvl, "level": lvl, "text": h.get("text") or ""})
    for r in covered_not_in_headings(page):
        rows.append({"seq": r["seq"], "kind": "pseudo",
                     "tag": "<%s>" % r["tag"], "level": int(r["suggested"][1]),
                     "text": r["text"], "suggested": r["suggested"]})
    rows.sort(key=lambda x: x["seq"])
    return rows


def structure_problems(page):
    """Deterministic structural problems for a page, from a crawler's view.
    Returns [(area, problem)]. Used as the baseline for the structure-fix section
    (the analyst may add richer, query-aware notes via gaps['structure_fix'])."""
    hc = page.get("heading_counts", {}) or {}
    h1, h2, h3 = hc.get("h1", 0), hc.get("h2", 0), hc.get("h3", 0)
    title = (page.get("title") or "").strip()
    out = []
    if h1 == 0:
        out.append(("H1", "No &lt;h1&gt; on the page — add exactly one H1 naming the primary topic."))
    elif h1 > 1:
        out.append(("H1", f"{h1} &lt;h1&gt; tags — keep exactly one; demote the extra(s) to H2."))
    if not title:
        out.append(("Title", "No &lt;title&gt; tag detected — add a concise, keyword-led title."))
    if h2 and h3 == 0:
        out.append(("Hierarchy", f"Flat outline — {h2} H2s and zero H3s. Sub-topics aren't nested; "
                                  "group detail points under their parent H2 as H3."))
    nh = covered_not_in_headings(page)
    if nh:
        out.append(("Heading markup", f"{len(nh)} topic(s) are styled as titles but use "
                    "&lt;p&gt;/&lt;div&gt;/&lt;span&gt;, not heading tags — invisible to crawlers "
                    "as headings (see promote list below)."))
    return out


def check_brand_keys(clusters, order):
    """Loud per-cluster warnings: a canonical brand that resolves to NO key in a
    cluster that DOES carry brand data — that column would render empty."""
    warnings = []
    for c in clusters.get("clusters", []):
        bmap = c.get("brands") or {}
        if not bmap:
            continue
        for b in order:
            if resolve_brand_key(bmap, b) is None:
                warnings.append(
                    "cluster %r: brand %r has NO matching key in clusters.json "
                    "(keys present: %s) -- its column would be empty."
                    % (c.get("name"), b, sorted(bmap.keys())))
    return warnings


def assert_your_column(clusters, your):
    """Fail loudly if the your-brand column is entirely zero across all clusters
    that carry brand data — almost always a key mismatch, not a real result."""
    cls = clusters.get("clusters", [])
    if not cls or not any((c.get("brands") or {}) for c in cls):
        return
    total = sum(int(brand_entry(c.get("brands"), your).get("depth", 0) or 0) for c in cls)
    if total == 0:
        bar = "=" * 72
        print("\n%s" % bar, file=sys.stderr)
        print("ERROR: the your-brand column (%r) is ZERO across all %d clusters."
              % (your, len(cls)), file=sys.stderr)
        print("This almost always means clusters.json keyed your page under a", file=sys.stderr)
        print("different string than meta.your_brand (e.g. 'OUR PAGE' or the URL)", file=sys.stderr)
        print("-- a key mismatch, not a real 'your page covers nothing' result.", file=sys.stderr)
        print("Fix: make every key in each cluster's 'brands' map exactly match", file=sys.stderr)
        print("meta.your_brand / competitors[].brand. The report was written but", file=sys.stderr)
        print("the your-brand column is INVALID until the keys are corrected.", file=sys.stderr)
        print("%s\n" % bar, file=sys.stderr)
        sys.exit(2)


# --------------------- content helpers (the new value) ----------------------
def best_section(page, cluster_name):
    """Pick the page section whose heading best matches the cluster name."""
    best, score = None, 0.0
    cn = (cluster_name or "").lower()
    for s in page.get("sections", []):
        r = difflib.SequenceMatcher(None, cn, (s.get("heading") or "").lower()).ratio()
        if r > score:
            best, score = s, r
    return best if score >= 0.34 else None


def similarity(a, b):
    a = (a or "")[:2500].lower()
    b = (b or "")[:2500].lower()
    if not a or not b:
        return 0.0
    return round(difflib.SequenceMatcher(None, a, b).ratio() * 100)


def sim_verdict(pct):
    if pct >= 80:
        return "near-duplicate"
    if pct >= 55:
        return "same idea, reworded"
    if pct >= 30:
        return "loosely related"
    return "distinct"


def textfrag(url, heading):
    """Deep-link that opens the live page scrolled to this heading (Chrome)."""
    if not url:
        return "#"
    frag = quote((heading or "")[:120])
    return f"{url}#:~:text={frag}" if frag else url


# ----------------------------- svg chart ------------------------------------
def svg_bar(title, pairs, unit=""):
    pairs = [(str(l), float(v or 0)) for l, v in pairs]
    if not pairs:
        return ""
    mx = max((v for _, v in pairs), default=0) or 1
    bar_h, gap, left, top, width = 22, 10, 170, 30, 540
    height = top + len(pairs) * (bar_h + gap) + 10
    rows = []
    for i, (label, val) in enumerate(pairs):
        y = top + i * (bar_h + gap)
        w = int((val / mx) * (width - left - 80))
        rows.append(f'<text x="0" y="{y+15}" class="lbl">{esc(label[:26])}</text>'
                    f'<rect x="{left}" y="{y}" width="{max(w,1)}" height="{bar_h}" rx="3" class="bar"/>'
                    f'<text x="{left+max(w,1)+6}" y="{y+15}" class="val">{val:g}{unit}</text>')
    return (f'<svg viewBox="0 0 {width} {height}" class="chart" role="img">'
            f'<text x="0" y="16" class="ctitle">{esc(title)}</text>' + "".join(rows) + "</svg>")


CSS = """
*{box-sizing:border-box} body{font-family:Segoe UI,Roboto,Helvetica,Arial,sans-serif;color:#0f172a;max-width:1180px;margin:0 auto;padding:16px 26px;line-height:1.5}
h1{font-size:24px;border-bottom:3px solid #1e3a8a;padding-bottom:6px;color:#1e3a8a}
h2{font-size:18px;margin-top:30px;border-bottom:1px solid #cbd5e1;padding-bottom:4px;color:#1e3a8a}
h3{font-size:15px;margin:14px 0 4px}
.kpis{display:flex;flex-wrap:wrap;gap:12px;margin:14px 0}
.kpi{flex:1 1 140px;background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;padding:12px 14px}
.kpi .n{font-size:24px;font-weight:700;color:#1e3a8a}.kpi .l{font-size:12px;color:#475569}
.banner{padding:10px 14px;border-radius:8px;margin:8px 0;font-size:14px}
.ok{background:#dcfce7;border:1px solid #16a34a;color:#166534}.warn{background:#fef3c7;border:1px solid #f59e0b;color:#92400e}
.toc{background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:10px 14px;font-size:13px}
.toc a{color:#1e40af;margin-right:14px;text-decoration:none}
table{border-collapse:collapse;width:100%;font-size:12.5px;margin:6px 0}
th,td{border:1px solid #cbd5e1;padding:4px 7px;text-align:left;vertical-align:top}
th{background:#eef2ff}.matrix td{text-align:center;font-weight:600}
.d0{background:#fee2e2;color:#991b1b}.d1{background:#fef3c7}.d2{background:#dbeafe}.d3{background:#dcfce7;color:#166534}
.cl{border:1px solid #e2e8f0;border-radius:10px;padding:10px 14px;margin:12px 0;background:#fbfdff}
.cl h3{color:#1e3a8a;margin-top:0}
.cards{display:flex;flex-wrap:wrap;gap:10px}
.card{flex:1 1 320px;border:1px solid #e2e8f0;border-left:4px solid #3b82f6;border-radius:8px;padding:8px 10px;background:#fff}
.card.ours{border-left-color:#16a34a;background:#f0fdf4}
.card .b{font-weight:700;font-size:13px}.card .ex{font-size:12px;color:#334155;margin-top:4px}
.badge{display:inline-block;font-size:10px;font-weight:700;padding:1px 6px;border-radius:4px;margin-left:4px}
.bd0{background:#fee2e2;color:#991b1b}.bd1{background:#fef3c7;color:#92400e}.bd2{background:#dbeafe;color:#1e40af}.bd3{background:#dcfce7;color:#166534}
.sim{font-size:12px;color:#475569;margin-top:6px}
.p3{color:#b91c1c;font-weight:700}.p2{color:#b45309;font-weight:600}.p1{color:#1d4ed8}
.chart{width:100%;max-width:560px}.chart .ctitle{font-size:13px;font-weight:700;fill:#1e3a8a}.chart .lbl{font-size:11px;fill:#334155}.chart .val{font-size:11px;fill:#475569}.chart .bar{fill:#3b82f6}
.charts{display:flex;flex-wrap:wrap;gap:24px}
.ol{font-size:13px;padding:1px 0}.lvl{display:inline-block;min-width:26px;font-size:10px;font-weight:700;color:#1e40af;font-family:monospace}
.mono{font-family:monospace;font-size:11px;word-break:break-all}
details{border:1px solid #e2e8f0;border-radius:8px;padding:8px 12px;margin:8px 0;background:#fbfdff}
summary{cursor:pointer;font-weight:600;color:#1e3a8a}small{color:#64748b}code{background:#f1f5f9;padding:1px 5px;border-radius:4px}
.controls input,.controls select{padding:6px 8px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px;margin-right:6px}
.struct{margin:6px 0 2px;border-left:2px solid #e2e8f0;padding-left:10px}
.srow{margin:3px 0}
.srow .tag{display:inline-block;min-width:30px;font-family:monospace;font-size:10px;font-weight:700;color:#fff;background:#1e3a8a;border-radius:4px;padding:1px 5px;text-align:center;margin-right:6px;vertical-align:middle}
.srow .sh{font-weight:700;color:#0f172a}
.srow .sc{color:#334155;margin:3px 0 9px;white-space:pre-wrap;font-size:12.5px}
.s1{margin-left:0}.s1 .sh{font-size:18px}
.s2{margin-left:16px}.s2 .sh{font-size:15px}
.s3{margin-left:34px}.s3 .sh{font-size:13.5px}
.s4{margin-left:52px}.s4 .sh{font-size:12.5px}
.s0 .sh{font-style:italic;color:#64748b;font-weight:600}.s0 .tag{background:#64748b}
.faq{border-left:3px solid #cbd5e1;padding:2px 0 2px 10px;margin:8px 0}
.faq .q{font-weight:700;color:#0f172a}.faq .a{color:#334155;font-size:12.5px;margin-top:2px;white-space:pre-wrap}
.hc td.yes{background:#dcfce7;color:#166534;text-align:center;font-weight:700}
.hc td.no{background:#fee2e2;color:#b91c1b;text-align:center}
.hc td:nth-child(-n+3){font-size:12px}.hc th{position:sticky;top:0}
@media print{details{break-inside:avoid}.cl{break-inside:avoid}.srow{break-inside:avoid}.faq{break-inside:avoid}}
"""
FILTER_JS = "function f(){var q=(document.getElementById('q').value||'').toLowerCase();var t=document.getElementById('ft').value,p=document.getElementById('fp').value;document.querySelectorAll('#gaps tbody tr').forEach(function(r){var okq=r.innerText.toLowerCase().indexOf(q)>=0;var okt=!t||r.dataset.type===t;var okp=!p||r.dataset.prio===p;r.style.display=(okq&&okt&&okp)?'':'none';});}"
# Expand every <details> when printing so the full page-structure (and other
# collapsible sections) render in the Print -> Save-as-PDF output, then restore.
PRINT_JS = ("window.addEventListener('beforeprint',function(){document.querySelectorAll('details').forEach(function(d){d.setAttribute('data-o',d.open?'1':'0');d.open=true;});});"
            "window.addEventListener('afterprint',function(){document.querySelectorAll('details').forEach(function(d){d.open=d.getAttribute('data-o')==='1';});});")


def outline_html(page):
    hc = page.get("heading_counts", {})
    cap = f'<small>H1:{hc.get("h1",0)} · H2:{hc.get("h2",0)} · H3:{hc.get("h3",0)} · H4+:{hc.get("h4_plus",0)} · words:{page.get("word_count_total",0)}</small>'
    rows = []
    for h in page.get("heading_outline", [])[:250]:
        lvl = int(h.get("level", 2) or 2)
        rows.append(f'<tr><td><b>H{lvl}</b></td><td style="padding-left:{(lvl-1)*16}px">{esc(h.get("text"))}</td></tr>')
    if not rows:
        return cap + " <small>— no headings captured</small>"
    return (cap + "<table><thead><tr><th style='width:56px'>Tag</th><th>Heading</th></tr></thead><tbody>"
            + "".join(rows) + "</tbody></table>")


def links_table(links, limit=60):
    if not links:
        return "<small>none captured</small>"
    out = ["<table><thead><tr><th>Anchor</th><th>Target</th><th>Section</th><th>Scope</th></tr></thead><tbody>"]
    for l in links[:limit]:
        out.append(f'<tr><td>{esc(l.get("anchor"))}</td><td class="mono">{esc(l.get("href"))}</td><td>{esc(l.get("section"))}</td><td>{esc(l.get("scope",""))}</td></tr>')
    out.append("</tbody></table>")
    if len(links) > limit:
        out.append(f"<small>… {len(links)-limit} more (see report.xlsx)</small>")
    return "".join(out)


def images_table(imgs, limit=40):
    if not imgs:
        return "<small>none captured</small>"
    out = ["<table><thead><tr><th>Alt text</th><th>Src</th></tr></thead><tbody>"]
    for im in imgs[:limit]:
        alt = im.get("alt") or "<span style='color:#b91c1c'>(no alt)</span>"
        out.append(f'<tr><td>{alt}</td><td class="mono">{esc(im.get("src"))}</td></tr>')
    out.append("</tbody></table>")
    return "".join(out)


def structure_html(page):
    """Replicate the page's published structure: every heading (with its H-tag
    and a level indent) followed by the copy beneath it, in document order."""
    rows = []
    for s in page.get("sections", []):
        lvl = int(s.get("level", 2) or 0)
        cls = "s%d" % (lvl if 0 <= lvl <= 4 else 4)
        tag = "intro" if lvl == 0 else "H%d" % lvl
        heading = esc(s.get("heading") or ("(intro / pre-heading copy)" if lvl == 0 else "(untitled)"))
        text = esc(s.get("text") or "")
        body = f'<div class="sc">{text}</div>' if text else '<div class="sc"><i>(no text captured)</i></div>'
        rows.append(f'<div class="srow {cls}"><span class="tag">{tag}</span>'
                    f'<span class="sh">{heading}</span> <small>· {esc(s.get("word_count"))} words</small>{body}</div>')
    if not rows:
        return "<small>no sections captured for this page</small>"
    return '<div class="struct">' + "".join(rows) + "</div>"


def build_html(run_dir, meta, clusters, gaps, pages, your, order):
    pbb = page_by_brand(pages)
    k = gaps.get("kpis", {})
    serp = gaps.get("serp", {})
    rank = gaps.get("ranking_assessment", {})
    ext = gaps.get("external_brands", {})
    topic = esc(gaps.get("topic") or meta.get("topic") or "this page")
    o = ["<!doctype html><html><head><meta charset='utf-8'>",
         f"<title>Content-Gap Report — {topic}</title><style>{CSS}</style></head><body>"]
    o.append(f"<h1>Content-Gap Report — {topic}</h1>")
    o.append(f"<small>Your page: <code>{esc(gaps.get('your_url') or meta.get('your_url'))}</code> · type: <b>{esc(gaps.get('page_type') or meta.get('page_type'))}</b> · pages compared: {len(order)}</small>")
    o.append("<div class='toc'><b>Jump to:</b> <a href='#kpi'>Overview</a><a href='#titles'>Titles &amp; H1</a>"
             "<a href='#matrix'>Cluster map</a><a href='#unique'>Unique coverage</a><a href='#coverage'>Coverage per company</a>"
             "<a href='#headcov'>Heading coverage</a><a href='#contentcov'>Content coverage</a>"
             "<a href='#content'>What each wrote (side-by-side)</a><a href='#struct'>Headings</a>"
             "<a href='#structure'>Page structure</a><a href='#faqs'>FAQs</a>"
             "<a href='#links'>Links</a><a href='#images'>Images</a><a href='#rank'>Ranking view</a>"
             "<a href='#gaps'>Gaps</a><a href='#structfix'>Structure fix</a></div>")

    yr = serp.get("your_rank")
    o.append(f"<div class='banner ok'>✅ Your page ranks <b>#{esc(yr)}</b> for \"{esc(serp.get('query'))}\".</div>"
             if yr else "<div class='banner warn'>⚠️ Your page did <b>not</b> appear in the top search results for this topic.</div>")

    o.append("<h2 id='kpi'>Overview</h2><div class='kpis'>")
    for key, label, u in [("coverage_pct", "Topic coverage", "%"), ("missing_count", "Missing", ""),
                          ("thin_count", "Thin", ""), ("faq_gap_count", "FAQ gaps", ""),
                          ("link_gap_count", "Link gaps", ""), ("quality_score", "Quality", "")]:
        o.append(f"<div class='kpi'><div class='n'>{esc(k.get(key,0))}{u}</div><div class='l'>{label}</div></div>")
    o.append("</div>")
    o.append("<div class='charts'>")
    o.append(svg_bar("Total words", [(p.get("brand", "?"), p.get("word_count_total", 0)) for p in pages]))
    o.append(svg_bar("H2 headings", [(p.get("brand", "?"), (p.get("heading_counts", {}) or {}).get("h2", 0)) for p in pages]))
    o.append(svg_bar("Internal links (on-page)", [(p.get("brand", "?"), len(onpage_internal_links(p))) for p in pages]))
    o.append("</div>")

    # pages analysed — page title + H1 per company
    o.append("<h2 id='titles'>Pages analysed — title &amp; H1</h2>")
    o.append("<small>The exact &lt;title&gt; and &lt;h1&gt; of each page compared, with its URL.</small>")
    o.append("<table><thead><tr><th>Company</th><th>Page title</th><th>H1</th><th>URL</th></tr></thead><tbody>")
    for b in order:
        p = pbb.get(b)
        if not p:
            continue
        o.append(f"<tr><td><b>{esc(b)}</b>{' (you)' if b == your else ''}</td>"
                 f"<td>{esc(p.get('title'))}</td><td>{esc(p.get('h1'))}</td>"
                 f"<td class='mono'><a href='{esc(p.get('url') or '#')}' target='_blank'>{esc(p.get('url'))}</a></td></tr>")
    o.append("</tbody></table>")

    # cluster matrix (diagram)
    o.append("<h2 id='matrix'>Cluster coverage map</h2><small>0 absent · 1 mention · 2 standard · 3 deep. Click a cluster to jump to the side-by-side content.</small>")
    o.append("<table class='matrix'><thead><tr><th style='text-align:left'>Topic cluster</th>")
    for b in order:
        o.append(f"<th>{esc(b)}</th>")
    o.append("</tr></thead><tbody>")
    for i, c in enumerate(clusters.get("clusters", [])):
        o.append(f"<tr><td style='text-align:left'><a href='#cl{i}'>{esc(c.get('name'))}</a></td>")
        for b in order:
            d = int(brand_entry(c.get("brands"), b).get("depth", 0) or 0)
            o.append(f"<td class='d{d}'>{d}</td>")
        o.append("</tr>")
    o.append("</tbody></table>")

    # missing from header tags — present on YOUR page, but no H1/H2/H3
    nh_your = covered_not_in_headings(pbb.get(your, {}) or {})
    o.append("<h3>Missing from header tags — on the page, but no H-tag</h3>")
    o.append("<small>Topics your page actually covers (rendered as styled titles) that carry "
             "<b>no</b> H1/H2/H3 tag, so they never register as headings during crawl. The cluster "
             "map above only counts real headings — these topics are invisible to it. Full "
             "per-company list under <a href='#headcov'>Heading coverage</a>.</small>")
    if nh_your:
        o.append("<ul>" + "".join(
            f"<li>{esc(r['text'])} <small>— currently &lt;{esc(r['tag'])}&gt;, suggest "
            f"<b>{r['suggested']}</b></small></li>" for r in nh_your[:40]) + "</ul>")
        if len(nh_your) > 40:
            o.append(f"<small>… and {len(nh_your)-40} more — see <a href='#headcov'>Heading coverage</a>.</small>")
    else:
        o.append("<small>none detected — every styled topic on your page is a real heading ✅</small>")

    # unique coverage — topics only ONE page covers
    uniq = unique_clusters_by_brand(clusters, order)
    o.append("<h2 id='unique'>Unique coverage — topics only one page covers</h2>")
    o.append("<small>Clusters that exactly one company addresses — a distinct angle the others miss. Yours = keep &amp; promote; a rival's = a gap you could close.</small>")
    o.append("<table><thead><tr><th>Company</th><th>Topics only this page covers</th></tr></thead><tbody>")
    for b in order:
        items = uniq.get(b) or []
        cell = ", ".join(esc(x) for x in items) if items else "<small>none</small>"
        o.append(f"<tr><td><b>{esc(b)}</b>{' (you)' if b == your else ''}</td><td>{cell}</td></tr>")
    o.append("</tbody></table>")

    # topic coverage per company — what each covers + a deep-link to that section
    o.append("<h2 id='coverage'>Topic coverage per company — with section links</h2>")
    o.append("<small>Per company: the topics they cover, how deep (0–3), the heading on their page, and a link that opens that exact section live. Click a company to expand.</small>")
    for b in order:
        p = pbb.get(b)
        if not p:
            continue
        rows = []
        for c in clusters.get("clusters", []):
            info = brand_entry(c.get("brands"), b)
            if not covers(info):
                continue
            d = int(info.get("depth", 0) or 0)
            sec = best_section(p, c.get("name"))
            head = (sec.get("heading") if sec else "") or info.get("section_heading") or ""
            dl = textfrag(p.get("url"), head) if head else (p.get("url") or "#")
            link = f"<a href='{esc(dl)}' target='_blank'>open ↗</a>" if head else "<small>—</small>"
            rows.append(f"<tr><td>{esc(c.get('name'))}</td><td class='d{d}' style='text-align:center'>{d}</td>"
                        f"<td>{esc(head)}</td><td>{link}</td></tr>")
        body = ("<table><thead><tr><th>Topic</th><th>Depth</th><th>Section heading on page</th><th>Link</th></tr></thead><tbody>"
                + "".join(rows) + "</tbody></table>") if rows else "<small>no covered topics captured for this page</small>"
        opn = " open" if p.get("is_ours") else ""
        o.append(f"<details{opn}><summary>{esc(b)}{' — OUR PAGE' if p.get('is_ours') else ''} · "
                 f"<small>{esc(p.get('title'))}</small></summary>{body}</details>")

    # heading coverage matrix — H1/H2/H3 x company (Yes/No), Sheet5-style
    hc_rows = heading_coverage(pages, order)
    o.append("<h2 id='headcov'>Heading coverage — H1 / H2 / H3 by company</h2>")
    o.append(f"<small>Every H1/H2/H3 heading across the pages ({len(hc_rows)} distinct), in <b>hierarchy order</b> — each H1 is followed by its H2s, each H2 by its H3s. The heading text sits in its level column; <b>Yes</b> = that company's page has this heading (matched case/punctuation-insensitively).</small>")
    o.append("<table class='hc'><thead><tr><th>H1</th><th>H2</th><th>H3</th>")
    for b in order:
        o.append(f"<th>{esc(b)}</th>")
    o.append("</tr></thead><tbody>")
    for r in hc_rows:
        cols = ["", "", ""]
        cols[r["level"] - 1] = esc(r["text"])
        o.append(f"<tr><td>{cols[0]}</td><td>{cols[1]}</td><td>{cols[2]}</td>")
        for b in order:
            yes = r["present"].get(b)
            o.append(f"<td class='{'yes' if yes else 'no'}'>{'Yes' if yes else 'No'}</td>")
        o.append("</tr>")
    o.append("</tbody></table>")

    # NEW: covered on the page but NOT in any heading tag (pseudo-headings)
    o.append("<h3>Covered on the page but NOT in any H1 / H2 / H3 tag</h3>")
    o.append("<small>Topics each page presents as a section title but ships inside a "
             "<code>&lt;p&gt; / &lt;div&gt; / &lt;span&gt;</code> (or a styled class) instead of a "
             "real heading tag — so a crawler / AI bot never sees them as headings. Promote each to "
             "the suggested tag. <b>Sits under</b> = the nearest real heading above it on the page.</small>")
    for b in order:
        p = pbb.get(b)
        if not p:
            continue
        rows = covered_not_in_headings(p)
        if not rows and not p.get("is_ours"):
            continue  # always show OUR page, even when clean
        opn = " open" if p.get("is_ours") else ""
        if rows:
            body = ("<table><thead><tr><th>Topic (styled as a heading)</th><th>Current tag</th>"
                    "<th>Sits under</th><th>Suggested fix</th></tr></thead><tbody>"
                    + "".join(
                        f"<tr><td>{esc(r['text'])}</td>"
                        f"<td class='mono'>&lt;{esc(r['tag'])}&gt;"
                        f"{(' .'+esc(r['class'])) if r['class'] else ''}</td>"
                        f"<td>{esc(r['parent_heading'])}</td>"
                        f"<td><b>promote to {r['suggested']}</b></td></tr>"
                        for r in rows[:200]) + "</tbody></table>")
            if len(rows) > 200:
                body += f"<small>… {len(rows)-200} more (see report.xlsx → ‘Covered Not In Headings’).</small>"
        else:
            body = "<small>none — every styled topic title on this page is a real heading tag ✅</small>"
        o.append(f"<details{opn}><summary>{esc(b)}{' — OUR PAGE' if p.get('is_ours') else ''} · "
                 f"{len(rows)} topic(s) not in heading tags</summary>{body}</details>")

    # content coverage matrix — topic x company (Yes/No), header-blind companion
    cc_rows = content_coverage(clusters, order)
    o.append("<h2 id='contentcov'>Content coverage — every topic, by company</h2>")
    o.append(f"<small>Every topic across the pages ({len(cc_rows)} distinct), <b>whether or not</b> it sits in a heading tag. "
             "<b>Yes</b> = that company's page covers this topic <b>anywhere</b> (heading or body copy) — even if it is "
             "worded differently (semantically matched by the gap analysis, so the same idea written in another company's "
             "words still counts). This is the header-blind companion to <a href='#headcov'>Heading coverage</a>: that table "
             "asks ‘is the topic an H-tag?’, this one asks ‘is the topic present at all?’.</small>")
    o.append("<table class='hc'><thead><tr><th style='text-align:left'>Topic</th>")
    for b in order:
        o.append(f"<th>{esc(b)}</th>")
    o.append("</tr></thead><tbody>")
    for r in cc_rows:
        o.append(f"<tr><td style='text-align:left'>{esc(r['topic'])}</td>")
        for b in order:
            yes = r["present"].get(b)
            o.append(f"<td class='{'yes' if yes else 'no'}'>{'Yes' if yes else 'No'}</td>")
        o.append("</tr>")
    o.append("</tbody></table>")

    # FAQ coverage matrix — question x company (Yes/No), differently-worded grouped
    fc_rows = faq_coverage(gaps, pages, order, your)
    o.append("<h3>FAQ coverage — each question, by company</h3>")
    o.append(f"<small>Every distinct FAQ across the pages ({len(fc_rows)} distinct). <b>Yes</b> = that company answers this "
             "question, even if they phrase it differently — differently-worded questions about the same thing are grouped "
             "into one row. Ranked most-covered first; the rows where your column reads <b>No</b> while others read Yes are "
             "your FAQ gaps.</small>")
    o.append("<table class='hc'><thead><tr><th style='text-align:left'>FAQ</th>")
    for b in order:
        o.append(f"<th>{esc(b)}</th>")
    o.append("</tr></thead><tbody>")
    for r in fc_rows:
        o.append(f"<tr><td style='text-align:left'>{esc(r['question'])}</td>")
        for b in order:
            yes = r["present"].get(b)
            o.append(f"<td class='{'yes' if yes else 'no'}'>{'Yes' if yes else 'No'}</td>")
        o.append("</tr>")
    o.append("</tbody></table>")

    # side-by-side: what each wrote, with similarity + deep links
    o.append("<h2 id='content'>What each page actually wrote — side by side</h2>")
    o.append("<small>Each card shows the real text the page wrote for that topic, its depth, a deep-link that opens that exact section on the live page, and how similar the wording is across pages.</small>")
    for i, c in enumerate(clusters.get("clusters", [])):
        o.append(f"<div class='cl' id='cl{i}'><h3>{esc(c.get('name'))}</h3><div class='cards'>")
        texts = {}
        for b in order:
            info = brand_entry(c.get("brands"), b)
            page = pbb.get(b, {})
            sec = best_section(page, c.get("name")) if info.get("present") else None
            txt = (sec.get("text") if sec else (info.get("snippet") or "")) or ""
            texts[b] = txt
            d = int(info.get("depth", 0) or 0)
            head = sec.get("heading") if sec else (info.get("section_heading") or "")
            dl = textfrag(page.get("url"), head) if head else page.get("url", "#")
            cls = "card ours" if b == your else "card"
            present = info.get("present")
            body = (esc(txt[:420]) + ("…" if len(txt) > 420 else "")) if present and txt else \
                ("<i>not covered on this page</i>" if not present else "<i>(no text captured)</i>")
            link = f" · <a href='{esc(dl)}' target='_blank'>open section ↗</a>" if head else ""
            o.append(f"<div class='{cls}'><div class='b'>{esc(b)}<span class='badge bd{d}'>depth {d}</span></div>"
                     f"<div class='ex'><b>{esc(head)}</b>{link}<br>{body}</div></div>")
        o.append("</div>")
        # similarity line
        present_brands = [b for b in order if texts.get(b)]
        sims = []
        for a in range(len(present_brands)):
            for bb in range(a + 1, len(present_brands)):
                pa, pbn = present_brands[a], present_brands[bb]
                pct = similarity(texts[pa], texts[pbn])
                sims.append(f"{esc(pa)}↔{esc(pbn)}: <b>{pct}%</b> ({sim_verdict(pct)})")
        if sims:
            avg = round(sum(int(s.split('<b>')[1].split('%')[0]) for s in sims) / len(sims))
            o.append(f"<div class='sim'><b>Content similarity:</b> {' · '.join(sims)} — "
                     f"overall ~{avg}% ({sim_verdict(avg)} wording across pages)</div>")
        o.append("</div>")

    # headings (separate H1/H2/H3 per page)
    o.append("<h2 id='struct'>Header tags — H1 / H2 / H3 per page</h2>")
    for p in pages:
        opn = " open" if p.get("is_ours") else ""
        o.append(f"<details{opn}><summary>{esc(p.get('brand'))}{' — OUR PAGE' if p.get('is_ours') else ''}</summary>{outline_html(p)}</details>")

    # full page structure — every heading + the content beneath it, in page order
    o.append("<h2 id='structure'>Full page structure — every heading + its content (page order)</h2>")
    o.append("<small>Each page reproduced as published: the H1/H2/H3 tag, the heading text, and the copy beneath it, in order — so you can see exactly where each heading sits on the live page. Click a page to expand.</small>")
    for b in order:
        p = pbb.get(b)
        if not p:
            continue
        opn = " open" if p.get("is_ours") else ""
        hc = p.get("heading_counts", {}) or {}
        cap = (f"H1:{hc.get('h1',0)} · H2:{hc.get('h2',0)} · H3:{hc.get('h3',0)} · "
               f"{len(p.get('sections', []))} sections · {p.get('word_count_total',0)} words")
        o.append(f"<details{opn}><summary>{esc(p.get('brand'))}{' — OUR PAGE' if p.get('is_ours') else ''} · "
                 f"<span class='mono'>{esc(p.get('url') or '')}</span> — {cap}</summary>{structure_html(p)}</details>")

    # FAQs per company — verbatim question + answer, segregated by page
    o.append("<h2 id='faqs'>FAQs by company — verbatim questions &amp; answers</h2>")
    o.append("<small>Every FAQ each page publishes, exactly as written — the question and how that company answered it. Compare phrasing and depth across competitors.</small>")
    for b in order:
        p = pbb.get(b)
        if not p:
            continue
        fqs = p.get("faqs") or []
        opn = " open" if p.get("is_ours") else ""
        if fqs:
            inner = "".join(
                f"<div class='faq'><div class='q'>Q: {esc(q.get('question'))}</div>"
                f"<div class='a'>{esc(q.get('answer')) or '<i>(no answer captured)</i>'}</div></div>"
                for q in fqs)
        else:
            inner = "<small>no FAQs captured on this page</small>"
        o.append(f"<details{opn}><summary>{esc(b)}{' — OUR PAGE' if p.get('is_ours') else ''} · {len(fqs)} FAQs</summary>{inner}</details>")

    # links — internal restricted to on-page (nav/header/footer excluded)
    o.append("<h2 id='links'>Internal &amp; external links per page</h2><small>Internal = <b>on-page editorial links only</b> — nav, header &amp; footer boilerplate excluded. Parsed from raw HTML.</small>")
    for p in pages:
        op = onpage_internal_links(p)
        o.append(f"<details><summary>{esc(p.get('brand'))} — {len(op)} on-page internal · {p.get('external_link_count',0)} external</summary>"
                 f"<h3>Internal (on-page)</h3>{links_table(op)}<h3>External</h3>{links_table(p.get('external_links', []))}</details>")

    # images
    o.append("<h2 id='images'>Images per page</h2>")
    for p in pages:
        o.append(f"<details><summary>{esc(p.get('brand'))} — {p.get('image_count',0)} images ({p.get('image_alt_count',0)} with alt)</summary>{images_table(p.get('images', []))}</details>")

    # ranking + external brands
    if rank or ext:
        o.append("<h2 id='rank'>Ranking view &amp; external mentions</h2>")
        if rank:
            o.append("<table><thead><tr><th>Page</th><th>Google search view</th><th>AI search / answer-engine view</th></tr></thead><tbody>")
            for b in order:
                r = map_get(rank, b, {}) or {}
                o.append(f"<tr><td>{esc(b)}</td><td>{esc(r.get('google'))}</td><td>{esc(r.get('ai_search'))}</td></tr>")
            o.append("</tbody></table>")
        if ext:
            o.append("<h3>External brands / third parties mentioned on each page</h3><ul>")
            for b in order:
                names = map_get(ext, b) or []
                o.append(f"<li><b>{esc(b)}</b>: {esc(', '.join(names)) if names else 'none detected'}</li>")
            o.append("</ul>")

    # gaps
    o.append("<h2 id='gaps'>Prioritised gaps — what to write</h2>")
    o.append("<div class='controls'><input id='q' onkeyup='f()' placeholder='search…'>"
             "<select id='ft' onchange='f()'><option value=''>all types</option><option>missing</option><option>thin</option><option>unique</option><option>faq</option><option>link</option><option>example</option><option>quality</option></select>"
             "<select id='fp' onchange='f()'><option value=''>all priorities</option><option value='3'>P3</option><option value='2'>P2</option><option value='1'>P1</option></select></div>")
    o.append("<table id='gaps'><thead><tr><th>P</th><th>Type</th><th>Cluster</th><th>What to do</th><th>Exemplar</th></tr></thead><tbody>")
    for g in sorted(gaps.get("gaps", []), key=lambda x: -int(x.get("priority", 0) or 0)):
        pr = int(g.get("priority", 0) or 0)
        o.append(f"<tr data-type='{esc(g.get('type'))}' data-prio='{pr}'><td class='p{pr}'>P{pr}</td><td>{esc(g.get('type'))}</td><td>{esc(g.get('cluster'))}</td>"
                 f"<td><b>{esc(g.get('title'))}</b><br><small>{esc(g.get('recommendation') or g.get('detail'))}</small></td><td>{esc(g.get('exemplar_brand'))}</td></tr>")
    o.append("</tbody></table>")

    # NEW: recommended structure fix — how a crawler / AI bot reads YOUR page,
    # and how to re-order / re-tag it. Prefers the analyst's query-aware notes
    # (gaps['structure_fix']); always backed by the deterministic baseline.
    ourp = pbb.get(your, {}) or {}
    sf = gaps.get("structure_fix") or {}
    o.append("<h2 id='structfix'>Recommended structure fix — how a crawler reads this page</h2>")
    o.append("<small>What a Google / AI crawl bot sees when it parses your page for the target query, "
             "and the structural changes that make the answer easier to find: which topics to promote "
             "to real headings, what the title / H1 should say, and the corrected reading order "
             "(what should sit higher, what lower). Based on your page: "
             f"<code>{esc(ourp.get('url') or '')}</code>.</small>")

    # problems (deterministic baseline + any analyst-supplied notes)
    probs = structure_problems(ourp)
    extra = sf.get("problems") or []
    o.append("<h3>What a crawler sees today — issues</h3>")
    if probs or extra:
        o.append("<ul>")
        for area, msg in probs:
            o.append(f"<li><b>{esc(area)}:</b> {msg}</li>")
        for m in extra:
            o.append(f"<li>{esc(m)}</li>")
        o.append("</ul>")
    else:
        o.append("<small>no structural issues detected ✅</small>")

    # title / H1 recommendation
    tr, hr1 = sf.get("title_recommendation"), sf.get("h1_recommendation")
    o.append("<table><thead><tr><th>Element</th><th>Currently on page</th><th>Recommendation</th></tr></thead><tbody>")
    o.append(f"<tr><td><b>&lt;title&gt;</b></td><td>{esc(ourp.get('title'))}</td>"
             f"<td>{esc(tr) if tr else '<small>keep — or tighten to lead with the target query</small>'}</td></tr>")
    o.append(f"<tr><td><b>H1</b></td><td>{esc(ourp.get('h1'))}</td>"
             f"<td>{esc(hr1) if hr1 else '<small>keep — exactly one H1 stating the primary topic</small>'}</td></tr>")
    o.append("</tbody></table>")

    # promote-to-heading list (the topics with no H-tag)
    promote = covered_not_in_headings(ourp)
    o.append("<h3>Promote these topics to real headings</h3>")
    if promote:
        o.append("<small>Each is on the page but in a non-heading tag. Promoting makes it a crawlable "
                 "section heading (and FAQ/section-snippet eligible).</small>")
        o.append("<table><thead><tr><th>Topic</th><th>Now</th><th>Sits under</th><th>Make it</th></tr></thead><tbody>")
        for r in promote[:80]:
            o.append(f"<tr><td>{esc(r['text'])}</td><td class='mono'>&lt;{esc(r['tag'])}&gt;</td>"
                     f"<td>{esc(r['parent_heading'])}</td><td><b>{r['suggested']}</b></td></tr>")
        o.append("</tbody></table>")
        if len(promote) > 80:
            o.append(f"<small>… {len(promote)-80} more in report.xlsx.</small>")
    else:
        o.append("<small>nothing to promote — all topics already use heading tags ✅</small>")

    # corrected reading order — real headings + promoted topics, in document order
    mo = merged_outline(ourp)
    o.append("<h3>Corrected reading order for crawlers</h3>")
    o.append("<small>Your page in document order with the missing topics slotted in where they "
             "physically appear — real headings shown with their tag, topics-to-promote flagged "
             "<span class='p2'>(promote)</span>. Use it to see what should sit higher vs lower and "
             "where each untagged topic belongs in the hierarchy.</small>")
    if mo:
        o.append("<div class='struct'>")
        for r in mo[:400]:
            lvl = int(r.get("level", 2) or 2)
            cls = "s%d" % (lvl if 1 <= lvl <= 4 else 4)
            flag = (f" <span class='p2'>(promote &lt;{esc(r['tag'])}&gt; → {r['suggested']})</span>"
                    if r["kind"] == "pseudo" else "")
            tag = esc(r["tag"]) if r["kind"] == "real" else r["suggested"]
            o.append(f"<div class='srow {cls}'><span class='tag'>{tag}</span>"
                     f"<span class='sh'>{esc(r['text'])}</span>{flag}</div>")
        o.append("</div>")
    else:
        o.append("<small>no outline captured for this page</small>")
    if sf.get("notes"):
        o.append(f"<p><small><b>Analyst note:</b> {esc(sf.get('notes'))}</small></p>")

    o.append("<hr><small>SEO Content-Gap Analyzer — identifies gaps & shows real competitor content for reference. Not publishable copy. Links/headings/images parsed from live HTML; verify before publishing.</small>")
    o.append(f"<script>{FILTER_JS}{PRINT_JS}</script></body></html>")
    path = os.path.join(run_dir, "report.html")
    with open(path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(o))
    return path


# ----------------------------- xlsx / csv ------------------------------------
def build_xlsx(run_dir, meta, clusters, gaps, pages, your, order):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.chart import BarChart, Reference
        from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
        from openpyxl.utils import get_column_letter
    except Exception:
        with open(os.path.join(run_dir, "full_content.csv"), "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh); w.writerow(["page", "level", "heading", "words", "text"])
            for p in pages:
                for s in p.get("sections", []):
                    w.writerow([p.get("brand"), "H" + str(s.get("level", "")), s.get("heading"), s.get("word_count"), s.get("text")])
        # Content / FAQ coverage Yes-No matrices as standalone CSVs too
        with open(os.path.join(run_dir, "content_coverage.csv"), "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh); w.writerow(["Topic"] + list(order))
            for r in content_coverage(clusters, order):
                w.writerow([r["topic"]] + ["Yes" if r["present"].get(b) else "No" for b in order])
        with open(os.path.join(run_dir, "faq_coverage.csv"), "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh); w.writerow(["FAQ"] + list(order))
            for r in faq_coverage(gaps, pages, order, your):
                w.writerow([r["question"]] + ["Yes" if r["present"].get(b) else "No" for b in order])
        return None, "openpyxl not installed — wrote full_content.csv + content_coverage.csv + faq_coverage.csv (pip install openpyxl for the full workbook)."

    pbb = page_by_brand(pages)
    wb = Workbook()
    head, fill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1E3A8A")
    wrap = Alignment(wrap_text=True, vertical="top")

    def sheet(name, cols):
        ws = wb.create_sheet(name)
        ws.append(cols)
        for c in ws[1]:
            c.font = head; c.fill = fill
        ws.freeze_panes = "A2"
        return ws

    # Overview
    ws = wb.active; ws.title = "Overview"
    ws.append(["KPI", "Value"])
    for c in ws[1]:
        c.font = head; c.fill = fill
    k = gaps.get("kpis", {})
    for label, key in [("Topic coverage %", "coverage_pct"), ("Missing", "missing_count"), ("Thin", "thin_count"),
                       ("FAQ gaps", "faq_gap_count"), ("Internal-link gaps", "link_gap_count"),
                       ("Quality score", "quality_score"), ("Your words", "your_word_count"),
                       ("Competitor median words", "competitor_median_word_count")]:
        ws.append([label, k.get(key, 0)])
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14

    # Charts — per-company metric bars (visual, mirrors the HTML charts)
    cws = sheet("Charts", ["Company", "Words", "Sections", "FAQs", "On-page internal links"])
    for b in order:
        p = pbb.get(b) or {}
        cws.append([b, p.get("word_count_total", 0), len(p.get("sections", []) or []),
                    len(p.get("faqs", []) or []), len(onpage_internal_links(p))])
    last = cws.max_row
    cws.column_dimensions["A"].width = 28
    if last > 1:
        for i, (title, col) in enumerate([("Total words", 2), ("Sections", 3),
                                          ("FAQs", 4), ("On-page internal links", 5)]):
            ch = BarChart(); ch.type = "bar"; ch.title = title; ch.legend = None
            ch.height = max(6, last * 0.55); ch.width = 20
            ch.add_data(Reference(cws, min_col=col, min_row=1, max_row=last), titles_from_data=True)
            ch.set_categories(Reference(cws, min_col=1, min_row=2, max_row=last))
            cws.add_chart(ch, "H%d" % (2 + i * 16))

    ws = sheet("Cluster Matrix", ["Cluster"] + order)
    for c in clusters.get("clusters", []):
        ws.append([c.get("name")] + [int(brand_entry(c.get("brands"), b).get("depth", 0) or 0) for b in order])
    ws.column_dimensions["A"].width = 38
    ws.auto_filter.ref = ws.dimensions
    if ws.max_row > 1 and order:
        rng = "B2:%s%d" % (get_column_letter(1 + len(order)), ws.max_row)
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="num", start_value=0, start_color="F8696B",
            mid_type="num", mid_value=1.5, mid_color="FFEB84",
            end_type="num", end_value=3, end_color="63BE7B"))

    # Page Titles & H1 — one row per company
    ws = sheet("Page Titles & H1", ["Company", "Title", "H1", "URL"])
    for b in order:
        p = pbb.get(b)
        if not p:
            continue
        ws.append([b, p.get("title"), p.get("h1"), p.get("url")])
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 40
    ws.auto_filter.ref = ws.dimensions

    # Unique Coverage — topics only one company covers
    ws = sheet("Unique Coverage", ["Company", "Topic only this page covers"])
    uniq = unique_clusters_by_brand(clusters, order)
    for b in order:
        for t in (uniq.get(b) or []):
            ws.append([b, t])
    ws.auto_filter.ref = ws.dimensions

    # Topic Coverage — per company: covered topics + deep-link to that section
    ws = sheet("Topic Coverage", ["Company", "Topic", "Depth", "Section heading", "Section link"])
    for b in order:
        p = pbb.get(b)
        if not p:
            continue
        for c in clusters.get("clusters", []):
            info = brand_entry(c.get("brands"), b)
            if not covers(info):
                continue
            sec = best_section(p, c.get("name"))
            head = (sec.get("heading") if sec else "") or info.get("section_heading") or ""
            link = textfrag(p.get("url"), head) if head else (p.get("url") or "")
            ws.append([b, c.get("name"), int(info.get("depth", 0) or 0), head, link])
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["E"].width = 60
    ws.auto_filter.ref = ws.dimensions

    # Heading Coverage — H1/H2/H3 x company (Yes/No) matrix, Sheet5-style
    ws = sheet("Heading Coverage", ["H1", "H2", "H3"] + order)
    for r in heading_coverage(pages, order):
        cells = ["", "", ""]
        cells[r["level"] - 1] = r["text"]
        ws.append(cells + ["Yes" if r["present"].get(b) else "No" for b in order])
    for col in ("A", "B", "C"):
        ws.column_dimensions[col].width = 34
    ws.auto_filter.ref = ws.dimensions
    if ws.max_row > 1 and order:
        rng = "D2:%s%d" % (get_column_letter(3 + len(order)), ws.max_row)
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="equal", formula=['"Yes"'], fill=PatternFill("solid", fgColor="DCFCE7")))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="equal", formula=['"No"'], fill=PatternFill("solid", fgColor="FEE2E2")))

    # Heading Markup Gaps — topics covered on each page but NOT in any H1/H2/H3
    ws = sheet("Heading Markup Gaps", ["Page", "Topic (styled as heading)", "Current tag", "Class", "Sits under", "Suggested"])
    for b in order:
        p = pbb.get(b)
        if not p:
            continue
        for r in covered_not_in_headings(p):
            ws.append([b, r["text"], "<%s>" % r["tag"], r["class"], r["parent_heading"], r["suggested"]])
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["E"].width = 36
    ws.auto_filter.ref = ws.dimensions

    # Structure Fix — issues + corrected reading order for OUR page
    ws = sheet("Structure Fix", ["Section", "Order", "Tag / Make", "Kind", "Detail"])
    ourp = pbb.get(your, {}) or {}
    for area, msg in structure_problems(ourp):
        ws.append(["ISSUE", "", area, "", re.sub(r"&[a-z]+;", "", msg)])
    tr, hr1 = (gaps.get("structure_fix") or {}).get("title_recommendation"), (gaps.get("structure_fix") or {}).get("h1_recommendation")
    ws.append(["TITLE", "", "<title>", "current", ourp.get("title")])
    if tr:
        ws.append(["TITLE", "", "<title>", "recommend", tr])
    ws.append(["H1", "", "H1", "current", ourp.get("h1")])
    if hr1:
        ws.append(["H1", "", "H1", "recommend", hr1])
    for r in merged_outline(ourp):
        tagcol = r["tag"] if r["kind"] == "real" else "%s → %s" % (r["tag"], r["suggested"])
        ws.append(["OUTLINE", r["seq"], tagcol, r["kind"], r["text"]])
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["E"].width = 70
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        row[0].alignment = wrap
    ws.auto_filter.ref = ws.dimensions

    # Content Coverage — topic x company (Yes/No), header-blind companion to Heading Coverage
    ws = sheet("Content Coverage", ["Topic"] + order)
    for r in content_coverage(clusters, order):
        ws.append([r["topic"]] + ["Yes" if r["present"].get(b) else "No" for b in order])
    ws.column_dimensions["A"].width = 46
    ws.auto_filter.ref = ws.dimensions
    if ws.max_row > 1 and order:
        rng = "B2:%s%d" % (get_column_letter(1 + len(order)), ws.max_row)
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="equal", formula=['"Yes"'], fill=PatternFill("solid", fgColor="DCFCE7")))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="equal", formula=['"No"'], fill=PatternFill("solid", fgColor="FEE2E2")))

    # FAQ Coverage — question x company (Yes/No), differently-worded questions grouped
    ws = sheet("FAQ Coverage", ["FAQ"] + order)
    for r in faq_coverage(gaps, pages, order, your):
        ws.append([r["question"]] + ["Yes" if r["present"].get(b) else "No" for b in order])
    ws.column_dimensions["A"].width = 64
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        row[0].alignment = wrap
    ws.auto_filter.ref = ws.dimensions
    if ws.max_row > 1 and order:
        rng = "B2:%s%d" % (get_column_letter(1 + len(order)), ws.max_row)
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="equal", formula=['"Yes"'], fill=PatternFill("solid", fgColor="DCFCE7")))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="equal", formula=['"No"'], fill=PatternFill("solid", fgColor="FEE2E2")))

    # Full Content — every section, every page
    ws = sheet("Full Content", ["Page", "Level", "Heading", "Words", "Text"])
    for p in pages:
        for s in p.get("sections", []):
            ws.append([p.get("brand"), "H" + str(s.get("level", "")), s.get("heading"), s.get("word_count"), s.get("text")])
    ws.column_dimensions["E"].width = 90
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        row[0].alignment = wrap
    ws.auto_filter.ref = ws.dimensions

    # Page Structure — published heading + its content in page order. One sheet;
    # filter the Page column to replicate any single site's full structure.
    ws = sheet("Page Structure", ["Page", "Order", "Tag", "Heading", "Words", "Content"])
    for b in order:
        p = pbb.get(b)
        if not p:
            continue
        for i, s in enumerate(p.get("sections", []), 1):
            lvl = int(s.get("level", 2) or 0)
            tag = "intro" if lvl == 0 else "H%d" % lvl
            ws.append([p.get("brand"), i, tag, s.get("heading"), s.get("word_count"), s.get("text")])
    ws.column_dimensions["D"].width = 42
    ws.column_dimensions["F"].width = 90
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        row[0].alignment = wrap
    ws.auto_filter.ref = ws.dimensions

    # Section Comparison — per cluster, each brand's text + similarity to ours
    ws = sheet("Section Comparison", ["Cluster", "Page", "Depth", "Section heading", "Words", "Similarity to OUR %", "Text"])
    for c in clusters.get("clusters", []):
        our_sec = best_section(pbb.get(your, {}), c.get("name"))
        our_txt = our_sec.get("text") if our_sec else ""
        for b in order:
            info = brand_entry(c.get("brands"), b)
            if not info.get("present"):
                continue
            sec = best_section(pbb.get(b, {}), c.get("name"))
            txt = sec.get("text") if sec else (info.get("snippet") or "")
            simp = "" if b == your else similarity(our_txt, txt)
            ws.append([c.get("name"), b, int(info.get("depth", 0) or 0),
                       sec.get("heading") if sec else "", sec.get("word_count") if sec else "", simp, txt])
    ws.column_dimensions["G"].width = 90
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        row[0].alignment = wrap
    ws.auto_filter.ref = ws.dimensions

    # Header tag lists — separate H1 / H2 / H3
    for lvl, nm in [(1, "H1 Tags"), (2, "H2 Tags"), (3, "H3 Tags")]:
        ws = sheet(nm, ["Page", "Heading"])
        for p in pages:
            for h in p.get("heading_outline", []):
                if int(h.get("level", 0) or 0) == lvl:
                    ws.append([p.get("brand"), h.get("text")])
        ws.auto_filter.ref = ws.dimensions

    # Internal links — on-page only (nav/header/footer boilerplate excluded)
    ws = sheet("Internal Links", ["Page", "Anchor", "Target", "Section", "Scope"])
    for p in pages:
        for l in onpage_internal_links(p):
            ws.append([p.get("brand"), l.get("anchor"), l.get("href"), l.get("section"), l.get("scope")])
    ws.auto_filter.ref = ws.dimensions

    ws = sheet("External Links", ["Page", "Anchor", "Target", "Section"])
    for p in pages:
        for l in p.get("external_links", []):
            ws.append([p.get("brand"), l.get("anchor"), l.get("href"), l.get("section")])
    ws.auto_filter.ref = ws.dimensions

    ws = sheet("Images", ["Page", "Alt text", "Src"])
    for p in pages:
        for im in p.get("images", []):
            ws.append([p.get("brand"), im.get("alt"), im.get("src")])
    ws.auto_filter.ref = ws.dimensions

    # Ranking + External brands (LLM-provided, optional)
    rank = gaps.get("ranking_assessment", {})
    if rank:
        ws = sheet("Ranking View", ["Page", "Google search view", "AI search view"])
        for b in order:
            r = map_get(rank, b, {}) or {}
            ws.append([b, r.get("google"), r.get("ai_search")])
        ws.auto_filter.ref = ws.dimensions
    ext = gaps.get("external_brands", {})
    if ext:
        ws = sheet("External Brands", ["Page", "Brands / third parties mentioned"])
        for b in order:
            ws.append([b, ", ".join(map_get(ext, b) or [])])
        ws.auto_filter.ref = ws.dimensions

    # Gaps / FAQ / Quality
    ws = sheet("Gaps", ["Priority", "Type", "Cluster", "Title", "Recommendation", "Exemplar"])
    prf = {3: PatternFill("solid", fgColor="FECACA"), 2: PatternFill("solid", fgColor="FEF3C7"), 1: PatternFill("solid", fgColor="DBEAFE")}
    for g in sorted(gaps.get("gaps", []), key=lambda x: -int(x.get("priority", 0) or 0)):
        ws.append([g.get("priority"), g.get("type"), g.get("cluster"), g.get("title"),
                   g.get("recommendation") or g.get("detail"), g.get("exemplar_brand")])
        f = prf.get(int(g.get("priority", 0) or 0))
        if f:
            ws.cell(row=ws.max_row, column=1).fill = f
    ws.auto_filter.ref = ws.dimensions

    ws = sheet("FAQ Gaps", ["Question", "Answered by"])
    for q in gaps.get("faq_gaps", []):
        ws.append([q.get("question"), ", ".join(q.get("answered_by", []))])
    ws.auto_filter.ref = ws.dimensions

    # FAQs by Company — verbatim Q&A, segregated per page
    ws = sheet("FAQs by Company", ["Company", "Question", "Answer"])
    for b in order:
        p = pbb.get(b)
        if not p:
            continue
        for q in (p.get("faqs") or []):
            ws.append([b, q.get("question"), q.get("answer")])
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 90
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
        for c in row:
            c.alignment = wrap
    ws.auto_filter.ref = ws.dimensions

    ws = sheet("Quality", ["Page", "Words", "H2s", "FAQs", "Internal links", "Schema", "E-E-A-T", "Freshness"])
    q = (gaps.get("quality") or {}).get("per_brand", {})
    for b in order:
        r = map_get(q, b, {}) or {}
        ws.append([b, r.get("word_count"), r.get("h2"), r.get("faqs"), r.get("internal_links"),
                   ", ".join(r.get("schema", []) or []), "yes" if r.get("eeat") else "no", r.get("freshness")])
    ws.auto_filter.ref = ws.dimensions

    path = os.path.join(run_dir, "report.xlsx")
    wb.save(path)
    return path, None


def main():
    if len(sys.argv) < 2:
        print("usage: python build_report.py <run_dir>"); sys.exit(1)
    run_dir = sys.argv[1]
    if not os.path.isdir(run_dir):
        print("run_dir not found:", run_dir); sys.exit(1)
    meta, clusters, gaps, pages = load_run(run_dir)
    your, order = brand_order(meta, gaps, pages)

    # Scrub form-widget junk (phone country-code runs) from every section's text
    # so it never appears in the structure / side-by-side / workbook.
    for p in pages:
        for s in (p.get("sections") or []):
            s["text"] = clean_text(s.get("text"))

    # Loud warnings BEFORE building — any canonical brand that won't resolve to a
    # key in a populated cluster, so an empty column can never ship unnoticed.
    for w in check_brand_keys(clusters, order):
        print("WARN :", w, file=sys.stderr)

    print("HTML  ->", build_html(run_dir, meta, clusters, gaps, pages, your, order))
    xlsx_path, note = build_xlsx(run_dir, meta, clusters, gaps, pages, your, order)
    if xlsx_path:
        print("XLSX  ->", xlsx_path)
    if note:
        print("NOTE :", note)
    print("Open report.html in a browser (Print -> Save as PDF to share).")

    # Post-build assertion: a fully-zero your-brand column is a key mismatch.
    assert_your_column(clusters, your)


if __name__ == "__main__":
    main()
